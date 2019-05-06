# -*- coding: utf-8 -*-
"""
Created on Thu Jul 26 14:41:52 2018

@author: bianl

修正：
1.将所有的输入项作为函数单独列出来，方便调试；
2.在选择课程项目的时候可以选择多个课程项目；
3.新教师x月产能，x默认值为用户第一次所选时间范围；
4.修复bug & 优化
"""
import datetime
import xlrd
import math
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import RED, BLUE


#  Set global variables
SCHEDULE_FILENMAE = "配课表明细16.6.1-18.5.31.xlsx"
ENTRYDATE_FILENAME = "国外部教师名单6.15.xlsx"
GAP_FILENAME = "离职&管理层老师名单.xlsx"
  
CLASSTYPE = {"英联邦VIP":["雅思","IELTS"],"雅思班级":["雅思","IELTS"],
                 "托福VIP":["TOEFL"],"托福班级":["TOEFL"],
                 "美本":["SSAT","SAT","ACT","AP","北美本科精英"],"美研":["GRE","GMAT"],
                 "其他":["企业培训班","英联邦中学","国际英语实验班","TOEIC","海外留学菁英"]} 



def LoadExcelTable( filename ):
    """
    载入配课表等excel文件
    filename为excel文件名
    return excel表头table
    """
    data = xlrd.open_workbook( filename )  
    table = data.sheets()[0]
    return table


def LoadTeachersGap( filename ):
    """
    载入离职&管理层老师名单, excel文件
    filename为文件名
    return 教师姓名list
    """
    #  打开excel
    data = xlrd.open_workbook( filename )  
    table = data.sheets()[0]
    return [table.row_values(i)[0] for i in range(table.nrows)]


def InputDateRange():
    """
    Function: 获取用户输入的日期范围(str)，直接回车则返回Default值None。
    Input: 无
    Output: datetime格式的日期范围，或Default值None
    """
    while True:       
        st = input("Please enter start date (e.g. 2015-12-21): ")
        ft = input("Please enter end date (e.g. 2015-12-21): ")
        #  Default case
        if len(st) < 1:
            stmp = None
            break
        else:
            try:
                stmp = datetime.datetime.strptime(st, "%Y-%m-%d")  # str转datetime
                break
            except:
                print("Please re-enter start date in this 2015-12-21 format. ")
                continue
    while True:
        #  Default case
        if len(ft) < 1:
            etmp = None
            break
        else:
            try:
                etmp = datetime.datetime.strptime(ft, "%Y-%m-%d")
                break
            except:
                print("Please re-enter start date in this 2015-12-21 format. ")
                continue
    return stmp, etmp
    

def GetDateRange( stmp, etmp, table ):
    """
    Function:根据用户输入的时间范围，确定table中的index范围
    
    Input:
    stmp, etmp: 用户输入的时间范围(datetime)
    table：载入的配课表数据
    
    return 起始与截止日期(datetime)，及其index, 该时间范围相当于多少个月
    """
    rows = table.nrows
    #  转换excel日期到datetime格式，并找到table中的最小和最大日期
    mindate = xlrd.xldate_as_datetime(table.row_values(1) [3],0)  
    maxdate = xlrd.xldate_as_datetime(table.row_values(rows - 1) [3],0)
    # Default value, 从第1行开始（第0行为header）
    if stmp is None or stmp <= mindate:
        start = 1
        p = mindate
    else:
        for i in range(1,rows - 1):
            curdate = xlrd.xldate_as_datetime(table.row_values(i) [3],0)
            if  curdate == stmp:
                start = i
                p = curdate
                break 
    # Default value, 在最后一行结束（Python下标从0开始，故最后一行index比excel小1）
    if etmp is None or etmp >= maxdate:
        end = rows - 1
        q = maxdate
    else:
        for i in range(max(2,start),rows - 1):# i至少从2开始
            curdate = xlrd.xldate_as_datetime(table.row_values(i) [3],0)
            predate = xlrd.xldate_as_datetime(table.row_values(i-1) [3],0)
            if predate <= etmp and curdate > etmp:
                end = i-1
                q = predate
                break
    day = (q-p).days
    month_rate = day / 30
    return p, q, start, end, round(month_rate,2)
                         

def GetConstraint():
    """
    没有input参数，通过屏幕获得用户输入参数
    返回Department Constraint 和 Class Type Constraint (str 或 None)
    """    
    #  输入要查询的部门
    #  Default为全部部门
    DeptType = ["北美项目部", "英联邦项目部"]
    while True:
        Dep = input("Please input department: '北美项目部', '英联邦项目部' or both(press Enter directly) ")   
        if Dep in DeptType or len(Dep) < 1:
            break
        else:
            print("Please re-enter department name")
                    
    #  输入要查询的课程项目
    #  提示所有的课程类型
    #  Default为全部类型    
    while True:
        print("共有7种课程项目，分别为：")  
        print([keys for keys in CLASSTYPE.keys()],end = " ")   
        # Class Type constraints
        cyconsts = input("Please input class types separated by spaces: (press Enter directly = select all the type) ") 
        if len(cyconsts) < 1:
            break
        flag = True
        cyconsts = cyconsts.split(" ")
        for i in range(len(cyconsts)):
            if cyconsts[i] not in CLASSTYPE.keys():
                print(cyconsts[i] + " is not a legal class type, Please re-enter class type name")
                flag = False
                break
        #  如果cyconsts是legal的，则终止循环
        if flag :
            break         
    return Dep, cyconsts


def isinClassType( Classprogram, CurrentClass, Capacity):
    """
    Function: 检查当前课程是否满足课程项目约束
    
    Input
    Classprogram: 课程项目约束(str)
    CurrentClass: 当前课程名(str)
    Capacity: 班级容量(str)
    
    Output
    True if 当前课程CurrentClass 属于 课程项目Classprogram; False otherwise
    """
    flag = False   
    for i in range(len(Classprogram)):
        if "VIP" in Classprogram[i]:  # VIP课一定是1对1 或 6人
            if Capacity != "1对1" and Capacity != "6人":
                continue
        if "班级" in Classprogram:  # 班级项目一定不是 1对1 或 6人的
            if Capacity == "1对1" or Capacity == "6人":
                continue             
        
        for j in range(len(CLASSTYPE[Classprogram[i]])):
            if CLASSTYPE[Classprogram[i]][j] in CurrentClass:
                flag = True
    return flag


def CheckConstraint(curdep, curcls, capacity, selfstudy, DepConstraint, CTConstraint):
    """
    Function: 分别检查“取消约束”，“自习约束”，“部门约束”，“课程约束”
    
    Input
    curdep: 当前部门
    curcls: 当前课程
    capacity: 班级容量
    selfstudy: 自习约束
    DepConstraint: 部门约束
    CTConstraint: 课程约束
    
    Output
    全部符合条件则return True, 否则return False
    """
    #  排除已取消的课程
    if "取消" in curcls:
        return False  
    #  排除自习课、辅导课、模考等
    if len(selfstudy) == 0:
        return False   
    #  Department Constraint
    if len(DepConstraint) >= 1:  # 没有使用default的情况
        if DepConstraint not in curdep:
            return False
    #  Class Type Constraint
    if len(CTConstraint) >=1:  # 没有使用default的情况
        return isinClassType(CTConstraint, curcls, capacity)
    #  部门 & 课程约束 选择Default的情况
    return True
                                    

def TableFilter( table, parameters ):
    """
    Function:梳理配课表table
        
    Input
    table: 载入的excel table
    parameter: [start_index, end_index, Department_Constraint, ClassType_Constraint]
    
    Output
    基于resclstab(matrix)，生成简化 & 统计后的配课表：ClassesTable_detail.txt（显示计算细节）
    classes: {课号:{教师姓名:[完成课时，完成课时比例，完成业绩]}} (dict)
    totalclsn: 总开班数目
    totaltn: 总参与教师数目
    """
    start = parameters[0]
    end = parameters[1]
    Dep = parameters[2]
    cyconsts = parameters[3]
    
    resclstab = []  # Result Class Table   
    classes = {}  # 用于计算的classes dictionary
    for i in range(start, end):
        #  逐行读取数据
        r = table.row_values(i)
        curdep = r[0]  # 当前部门
        curcls = r[2]  # 当前课程种类  
        clsnum = r[1]  # 当前课号
        cur_t = r[5]  # 当前教师            
        #  Check该行数据是否满足约束
        flag = CheckConstraint(curdep, curcls, r[7], cur_t, Dep, cyconsts)                      
        if  not flag:
            continue             
        if clsnum not in classes.keys():
            #  当前课号第一次出现
            t4tc = {cur_t:1}  # teacher for this class
            #  更新classes dictionary
            classes[clsnum] = t4tc      
            #  更新Result Class Table   
            tmp = [curdep,curcls,clsnum,r[4],r[10],r[7],r[11]]
            resclstab.append(tmp)
        else:
            #  当前课号已存在于classes中
            if cur_t in classes[clsnum].keys():  # 当前教师已存在于该课号中
                classes[clsnum][cur_t] += 1
            else:  # 当前教师不存在于该课号中
                classes[clsnum][cur_t] = 1                   
    #  修改classes和resclstab  
    totalclsn = len(resclstab)  # Total classes number      
    teacherlist = {}  #  参与教师dict初值     
    for i in range(totalclsn):
        clsnum = resclstab[i][2]
        #  计算该班的业绩
        if isinstance(resclstab[i][6],float) :
            if resclstab[i][6] > 0:
                fee = resclstab[i][3] * resclstab[i][6]  # 学费乘以人数
            else: 
                if resclstab[i][5] == "6人":
                    fee = resclstab[i][3] * 6
                else: 
                    fee = resclstab[i][3]
        else:
            fee = 0
        tct = resclstab[i][4]  # Total Class Times 
        #  更新该课程中每位老师的课数、课时贡献百分比 和 对应的完成业绩
        for t in classes[clsnum].keys(): 
            if t not in teacherlist.keys():
                teacherlist[t] = 0
            times = classes[clsnum][t]
            p = times / tct
            classes[clsnum][t] = [times,round(p,2),round(p*fee,2)]
        resclstab[i].append(classes[clsnum])
    totaltn = len(teacherlist.keys())  # 参与教师总数
    #  输出文件：“ClassesTable_detail.txt”
    f = open("ClassesTable_detail.txt" , "w+")
    for i in range(len(resclstab)):
        f.writelines(str(resclstab[i]) + "\n")
    f.close()
    return classes, totalclsn, totaltn,

             
def GetGap ( classes, gap_list ):
    """
    Input
    classes: {课号:{教师姓名:[完成课时，完成课时比例，完成业绩]}} (dict)
    
    Output
    availableT: available teacher list，在职教师名及其完成业绩(tuple)
    unavailableT: unavailable teacher list，离职 & 管理岗教师名及其完成业绩(tuple)
    gap:由离职 & 管理岗教师产生的业绩缺口(float)
    """
    gap = 0
    availableT = []
    unavailableT = []
    teacherlist = {}
    #  计算教师产值列表
    for c in classes.keys():
         for t in classes[c].keys():                       
             individ = classes[c][t][2]  # 个人业绩
             if t not in teacherlist.keys():
                 teacherlist[t] = round(individ,2)
             else:
                 teacherlist[t] += round(individ,2)
    #  排序教师产值
    #  对teacherlist（dict）的value排序
    tmp = sorted(teacherlist.items(),key = lambda item:item[1], reverse = True)    
    #  从availableT中找出unavailable的教师
    for i in range(len(tmp)):
        if tmp[i][0] in gap_list:
            gap += tmp[i][1]
            unavailableT.append(tmp[i])
        else:
            availableT.append(tmp[i])
    return availableT, unavailableT, gap


def NewTeacherDateRange( timerate ):
    """
    Function: 通过用户输入，得到预估新教师产能的日期范围
    
    Input
    默认值为用户在outer loop中输入的日期范围，timerate
    
    Return duration，新教师入职x个月的产能表现。
    """  
    #  考虑3个月培训期
    timerate += 3  
    prompt = ("新教师入职x个月的产能表现，请输入x值(Deafault x = "
              + str(timerate) + " )：")   
    duration = input(prompt)
    if len(duration) < 1:  # Default case
        return timerate
    else:
        return float(duration)
    
    
def GetDateRange4NTP ( sche_table, obdate, enddate ):
    """
    Get Date Range for New Teachers's Performance
    Input
    sche_table: 配课表table
    startdate - enddate: 表示新教师入职后的测试时间段
        
    Output
    Index:[start(int), end(int), month_rate(float)] (list)
        即[起始日期的index, 终止日期的index，在此期间该教师工作了多少个月](list)
    """
    rows = sche_table.nrows
    #  转换excel日期到datetime格式，并找到sche_table中的最小和最大日期
    #  mindate - maxdate: 表示班表中的时间范围
    mindate = xlrd.xldate_as_datetime(sche_table.row_values(1) [3],0)  
    maxdate = xlrd.xldate_as_datetime(sche_table.row_values(rows - 1) [3],0)   
    #  (s,e)[min,max] 或 [min,max](s,e)的情况
    if enddate < mindate or obdate > maxdate:
        return 
    #  [min (s,e) max]的情况
    elif obdate > mindate and enddate < maxdate:
        for i in range(2,rows - 1):  
            curdate = xlrd.xldate_as_datetime(sche_table.row_values(i) [3],0)
            predate = xlrd.xldate_as_datetime(sche_table.row_values(i-1) [3],0)
            if  curdate == obdate:
                start = i
                p = curdate
            if predate <= enddate and curdate > enddate:
                end = i-1   
                q = predate
                break
    #  (s ==[min, e) max]的情况
    elif obdate <= mindate and enddate < maxdate:
        start = 1
        p = mindate
        for i in range(2,rows - 1):  
            curdate = xlrd.xldate_as_datetime(sche_table.row_values(i) [3],0)
            predate = xlrd.xldate_as_datetime(sche_table.row_values(i-1) [3],0)
            if predate <= enddate and curdate > enddate:
                end = i-1
                q = predate
                break
    #  [min (s, max] == e)的情况
    elif obdate > mindate and enddate >= maxdate:  
        end = rows - 1
        q = maxdate              
        for i in range(2,rows - 1):      
            curdate = xlrd.xldate_as_datetime(sche_table.row_values(i) [3],0)
            if  curdate == obdate:
                start = i
                p = curdate
                break  
    #  p和q表示最终计算的日期范围
    #  受培训期影响，p通常比obdate晚3个月
    #  month_rate为[p,q]期间工作的月份
    day = (q-p).days
    month_rate = day / 30
    return [start, end, month_rate]


def CalNTP( table, name, Dep, cyconsts, Index ):  
    """
    Function: Calculate a New Teacher's Performance per month
    
    Input
    table: 配课表信息
    name: 教师姓名
    Dep: 部门约束
    cyconsts: 课程约束
    Index:[起始日期的index(int), 终止日期的index(int)，在此期间该教师工作了多少个月(float)](list)
            
    Output
    name老师入职Index[0] - Index[1]时间内，完成的平均月业绩    
    """
    start = Index[0]
    end = Index[1]
    month_rate = Index[2]
    
    clsdict = {}  # 表示老师所授所有课程，及其在其中所贡献业绩的classes dictionary
    #  读入excel   
    for i in range(start, end):
        r = table.row_values(i)   
        #  教师姓名约束
        if name not in r[5]:
            continue   
        #  约束检查
        flag = CheckConstraint(r[0], r[2], r[7], r[5], Dep, cyconsts)
        if not flag:
            continue                       
        #  计算该班的业绩
        if isinstance(r[11],float) :
            if r[11] > 0:
                fee = r[4] * r[11]  # 学费乘以人数
            else: 
                if r[7] == "6人":
                    fee = r[4] * 6
                else: 
                    fee = r[4]
        else:
            fee = 0
        #  计算classes dictionary      
        if r[1] not in clsdict.keys():  # 首次记录该课号
            clsdict[r[1]] = (1/r[10]) * fee
        else:
            clsdict[r[1]] += (1/r[10]) * fee
    sum = 0
    for item in clsdict.keys():
        sum += clsdict[item]
    #  计算月平均业绩
    #  sum为0，或month_rate为负表示仍在培训期内，返回0
    return round(max(0,sum/month_rate),2)
          

def NewTeacherPerform( sche_table, entry_table, duration, Dep, cyconsts ):
    """
    Function: 计算符合约束条件的所有教师们，在入职一段时间后的月平均完成业绩表现
    
    Input
    sche_table: 配课表信息
    entry_table：教师入职时间表信息
    Dep: 部门约束
    cyconsts: 课程项目约束
    
    Output
    NTPerform(list): 符合条件的老师的月平均完成业绩表现
    performMean(float): “新”教师月平均表现的平均值
    """
    perform = {}
    m = entry_table.nrows 
    for i in range(1, m-1):
        r = entry_table.row_values(i) 
        name = r[2]  # teacher's name      
        obdate = xlrd.xldate_as_datetime(r[3],0)  #  Onboarding date
        enddate = obdate + datetime.timedelta(days = (int(duration) * 30))
        Index = GetDateRange4NTP(sche_table, obdate, enddate)
        if Index is None:
            continue
        score = CalNTP(sche_table, name, Dep, cyconsts, Index)
        perform[name] = score   
    #  对perform（dict）的value排序    
    tmp = sorted(perform.items(),key = lambda item:item[1], reverse = True) 
    #  删除还在培训期的老师
#    tmp = [NTPerform[i][1] for i in range(len(NTPerform)) if NTPerform[i][1] != 0]
#    tmp = np.array(tmp)    
    NTPerform =  [tmp[i][:] for i in range(len(tmp)) if tmp[i][1] != 0]
    tmp = np.array([NTPerform[i][1] for i in range(len(NTPerform))])
    return NTPerform, round(np.mean(tmp),2) 


def InputGoalProfit():
    """
    """
    #  输入目标业绩
    while True:
        profit = input("请输入（真实）业绩目标（万元）: ")   
        try:
            profit = float(profit) * 10000
            break
        except:
            print("请重新输入数字")
    return profit


def NumofTneeded( availableT, gap, estimean, goal ):
    """
    Function: Calculate Number of Teachers needed
    
    Input
    availableT(list): 可用教师的产能列表
    gap(float): 由不可用教师产生的产能缺口
    estimean(float): 新教师的预估产能
    goal(float): 目标面上业绩
    
    Output
    totalfee(float): 当前能够承担的总产能
    totalgap(float): 面上产能总缺口
    NTnum(int): 需要招聘教师数
    """

    #  现有可用教师能够承担的总业绩
    totalfee = 0
    for i in range(len(availableT)):
        totalfee += availableT[i][1]  
    #  总的产能缺口
    totalgap = goal - totalfee
    #  需要的新教师数
    NTnum =  math.ceil(totalgap / estimean)
    return totalfee, totalgap, NTnum


def FindZombie( availableT, estimean ):
    """
    Function: 找到可用教师产能列表中，产能小于等于新教师平均产能的行[教师，产能]
    
    Input
    availableT: 可用教师产能列表    
    estimean: 新教师平均产能
    
    Return
    僵尸教师的index起始值与终止值    
    """
    n = len(availableT)
    for i in range(n):
        if availableT[i][1] <= estimean:
            return i, n
        
    
def Output( startdate, enddate, timerate, Dep, cyconsts, 
            totalclsn, totaltn, totalfee, goal, 
            availableT, unavailableT, gap, totalgap,
            NTPerform, performMean, duration, NTnum ):
    """
    Function: 将4部分信息，输出到一个excel文件中
    startdate, enddate: 分析数据的起始 & 终止日期
    
    Output
    Part1: 可用教师完成业绩情况
    Part2: 不可用教师完成业绩情况
    Part3: 新教师表现
    Part4: 总结，需要招聘人数
    """
    estimean = timerate * performMean
    daterange = startdate.strftime("%Y-%m-%d") + "至" + enddate.strftime("%Y-%m-%d")
    filename = daterange + "招聘人数说明.xlsx"
    
    try:
        #  文件已存在的话，打开文件
        wb = load_workbook( filename ) 
        #  新建一个sheet
        ws = wb.create_sheet() 
    except:
        #  文件不存在，首次写入文件
        #  在内存中创建一个workbook对象，而自动创建一个worksheet   
        wb = Workbook()   
        #  获取当前活跃的worksheet，默认就是第一个worksheet
        ws = wb.active 
    finally:      
        #  Part1部分   
        if len(Dep) < 1:
            department = "全部部门"
        if len(cyconsts) < 1:
            program = "全部课程项目"
        else:
            program = '，'.join(cyconsts) + "课程项目"   
        part1strA4 = "可用教师完成业绩：" + str(round(totalfee,2)) + "元"
        part1strD2 = "共开班: " + str(totalclsn) + "节"
        part1strD3 = "共" + str(totaltn) + "位教师参与教学"
        #  部分字体格式  
        font1 = Font( name = 'Calibri', size = 11, bold = True, italic = False, vertAlign = None,
                     underline = 'none', strike = False, color = RED )  
        #  Part1内容
        ws["A1"].font = ws["A2"].font = ws["A3"].font = ws["A4"].font = ws["D2"].font = ws["D3"].font = font1    
        ws["A1"] = daterange; ws["A2"] = department; ws["A3"] = program; ws["A4"] = part1strA4       
        for i in range(len(availableT)):
            ws.append(availableT[i])
        ws["D2"] = part1strD2; ws["D3"] = part1strD3
        #  将僵尸教师字体改为蓝色
        [p, r] = FindZombie( availableT, estimean )
        p += 5; r += 5
        for i in range(p, r):
            ws.cell(row = i, column = 1).font = Font( color = BLUE ) 
            ws.cell(row = i, column = 2).font = Font( color = BLUE ) 
        ws.cell(row = p, column = 3, value = "← 小于等于新教师平均表现").font = Font( bold = True, color = RED ) 
        #  Part2内容
        ws["F1"].font = ws["F2"].font = ws["F3"].font = ws["F4"].font = font1
        part2F4 = "离职 & 管理层老师形成的业绩缺口：" + str(round(gap,2)) + "元"
        ws["F1"] = daterange; ws["F2"] = department; ws["F3"] = program; ws["F4"] = part2F4
        for i in range(len(unavailableT)):
            ws.cell(row = 5 + i, column = 6 , value = unavailableT[i][0])
            ws.cell(row = 5 + i, column = 7 , value = unavailableT[i][1])
        #  Part3内容    
        part3K2 = "教师入职满" + str(duration) + "个月时的产能月平均表现："
        part3K3 = "新教师产能月平均表现的均值为：" + str(round(performMean,2)) + "元"
        part3K4 = "对应" + daterange + ", " + str(round(timerate,2)) + "个月的产能：" + str(round(estimean,2)) + "元" 
        ws["K1"].font = ws["K2"].font = ws["K3"].font = ws["K4"].font = ws["L1"].font = font1
        ws["K1"] = department; ws["L1"] = program; ws["K2"] = part3K2; ws["K3"] = part3K3
        ws["K4"] = part3K4
        for i in range(len(NTPerform)):
            ws.cell(row = 5 + i, column = 11 , value = NTPerform[i][0])
            ws.cell(row = 5 + i, column = 12 , value = NTPerform[i][1])
        #  Part4内容
        part4N8 = ("目标产能缺口为：" + str(round(goal - totalfee - gap , 2)) + 
                   "元，不可用教师缺口为：" + str(round(gap,2)) + "元" +
                   "面上产能总缺口为：" + str(round(totalgap, 2)) + "元")
        part4N9 = '''如果维持"不活跃"教师现状不变，则此段时间所需的新教师数：'''
        part4N10 = "等于 面上产能总缺口 / 新教师平均产能表现 "
        part4N11 = "等于：" + str(NTnum) + "人"
        ws["N7"].font = ws["N8"].font = ws["N9"].font = ws["N10"].font = ws["N11"].font = font1
        ws["N14"].font = Font( size = 10, color = BLUE )
        ws["N7"] = "Summary: "; ws["N8"] = part4N8; ws["N9"] = part4N9; ws["N10"] = part4N10
        ws["N11"] = part4N11       
        #  保存文件
        wb.save( filename )


def DoesChangeDep():
    """
    Function: 是否重新选择部门与项目进行下一次分析
    Return True如果需要重新分析，False otherwise
    """
    while True:        
        ans = input("Choose another department/program to analysis again ? [y / n]: ")
        if ans != "n" and ans != "y":
            print("Please input the legal parameter [y / n]")
            continue
        if ans == "y":
            return True
        elif ans == "n":
            return False
        

def DoesChangeDate():  
    """
    Function: 是否重新选择时间范围进行下一次分析
    Return True如果需要重新分析，False otherwise
    """
    while True:        
        ans = input("Choose another date range to analysis again ? [y / n]: ")
        if ans != "n" and ans != "y":
            print("Please input the legal parameter [y / n]")
            continue
        if ans == "y":
            return True
        elif ans == "n":
            return False

              
def CalcuNum( sche_table, entry_table, gap_list ):
    while True:       
        # 用户输入 & 转换日期范围
        [stmp, etmp] = InputDateRange()
        [startdate, enddate, start, end, timerate] = GetDateRange( stmp, etmp, sche_table )                 
        while True:
            [Dep, cyconsts] = GetConstraint()  #  部门约束 & 课程约束                                    
            #  Step1.计算约束条件下的教师产能列表，缺口等统计数据
            #  梳理 & 统计 每个班对应的各教师贡献业绩，总开课数，总教师数
            [classes, totalclsn, totaltn] = TableFilter( sche_table, [start, end, Dep, cyconsts] )                     
            #  计算可用 & 非可用教师名单，以及由此产生的业绩缺口            
            [availableT, unavailableT, gap] = GetGap( classes, gap_list )                    
            #  Step2. 计算老师入职后一段时间的平均表现
            duration = NewTeacherDateRange( timerate )           
            [NTPerform, performMean] = NewTeacherPerform( sche_table, entry_table, duration, Dep, cyconsts )
            #  Step3. 计算需要招聘人数
            estimean = timerate * performMean  #  新教师在所选日期范围内的平均业绩表现
            goal = InputGoalProfit()  #  目标业绩
            [totalfee, totalgap, NTnum] = NumofTneeded( availableT, gap, estimean, goal )
            #  Step4. 输出计算结果
            Output( startdate, enddate, timerate, Dep, cyconsts, 
                    totalclsn, totaltn, totalfee, goal, 
                    availableT, unavailableT, gap, totalgap,
                    NTPerform, performMean, duration, NTnum )
            #  是否重新选择部门 / 项目分析
            if not DoesChangeDep():
                break
        #  是否重新选择时间分析
        if not DoesChangeDate():
            break

        
#  Main program
if __name__ == '__main__':
    #  载入配课表, 教师入职时间表, 离职&管理层老师名单
    sche_table = LoadExcelTable( SCHEDULE_FILENMAE ) 
    entry_table = LoadExcelTable( ENTRYDATE_FILENAME )
    gap_list = LoadTeachersGap( GAP_FILENAME ) 
    #  计算缺口 & 招聘人数  
    CalcuNum( sche_table, entry_table, gap_list )