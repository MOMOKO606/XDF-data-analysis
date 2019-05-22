import datetime
import os
import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook


class TpTrackers():
    """
    定义班表追踪器类，Teachers' performance Trackers。
    功能：提取并整理excel格式班表中的关键信息。
    """


    def __init__( self, CLASS_SCHEDULE_FILE ):
        """
        function: 初始化班表追踪器。
        :param CLASS_SCHEDULE_FILE: 班表文件名。
        """

        #  打开excel。
        #  table, rows, cols保存为类属性。
        data = xlrd.open_workbook( CLASS_SCHEDULE_FILE )
        table = data.sheets()[0]
        self.table = table
        self.rows = table.nrows
        self.cols = table.ncols

        #  需要用到的一些常量属性。
        self.DEPT_TYPE = ["北美项目部", "英联邦项目部"]  # 所有的项目部门。
        #  所有的课程类型。
        self.CLASS_TYPE = {"英联邦VIP": ["雅思", "IELTS", "GCSE"], "雅思班级": ["雅思", "IELTS"],
                      "托福VIP": ["TOEFL"], "托福班级": ["TOEFL"],
                      "美研": ["GRE", "GMAT"], "美本": ["SSAT", "SAT", "ACT", "AP", "北美本科精英"],
                      "其他": ["企业培训班", "英联邦中学", "国际英语实验班", "TOEIC", "海外留学菁英"]}
        self.DEPT_POS = 17  # 部门信息位于excel中的第几列（从0开始），例如，D列即为3，第3列。
        self.CLASS_NUM = 0  # 课号信息（班级编码）位于excel中的第几列（从0开始）。
        self.CLASS_POS = 1  # 课程种类信息位于excel中的第几列（从0开始）。
        self.CLASS_DATE_POS = 3  # 上课日期位于excel中的第几列（从0开始）
        self.FEE_POS = 2  # 学费信息位于excel中的第几列（从0开始）。
        self.TEACHERS_POS = 11  # 教师姓名信息位于excel中的第几列（从0开始）。
        self.OPENDATE_POS = 21  # 开班日期信息位于excel中的第几列（从0开始）。
        self.CLOSEDATE_POS = 22  # 结课日期信息位于excel中的第几列（从0开始）。
        self.CLASS_TIME_POS = 6  # 本次课时（分钟）信息位于excel中的第几列（从0开始）。
        self.CLASS_TIMESUM_POS = 24  # 总课时(分钟)信息位于excel中的第几列（从0开始）。
        self.STU_NUM_POS = 29  # 当前学生数量信息位于excel中的第几列（从0开始）。


    def GetDateLimit( self ):
        """
        function: 班表中的上课日期可能不是排序好的，此时需要遍历班表，获取最小和最大日期。
        :return: mindate & maxdate，即此班表中的日期范围。
        """

        #  载入参数。
        table = self.table
        CLASS_DATE_POS =self.CLASS_DATE_POS

        #  赋初值。
        mindate = xlrd.xldate_as_datetime(table.row_values(1)[CLASS_DATE_POS], 0)
        maxdate = xlrd.xldate_as_datetime(table.row_values(1)[CLASS_DATE_POS], 0)

        #  excel班表为1到m行，读入table后自动转为0到m-1行。
        #  第0行为header，第1行赋初值，所以从第2行循环至第m-1行。
        for i in range(2, self.rows):
            #  每个班上课的日期。
            tmpdate = xlrd.xldate_as_datetime(table.row_values(i)[CLASS_DATE_POS], 0)
            if tmpdate < mindate:
                mindate = tmpdate
            if tmpdate > maxdate:
                maxdate = tmpdate
        return mindate, maxdate


    def GetDateRange( self, st, ft ):
        """
        function: 根据用户输入的起始日期和班表中的日期，计算所需的日期范围。
        :param st: ft:查询的起始与终止日期，Default value为空，即班表中的全部日期。
        :return: 日期范围startdate, enddate。
        """

        #  转换excel日期到datetime格式，并找到table中的最小和最大日期。
        #  注意，excel中的日期格式应为2018/10/22形式，其他形式需在excel中先进行预处理。
        [mindate, maxdate] = self.GetDateLimit()

        #  根据用户输入的日期范围，确定最终的日期范围。

        if len(st) < 1:  # Default value, 从第1行开始（第0行为header）
            startdate = mindate
        else:
            try:
                stmp = datetime.datetime.strptime(st, "%Y-%m-%d")  # str转datetime
                if stmp <= mindate:
                    startdate = mindate
                else:
                    startdate = stmp
            except:
                print("Wrong start date format, please rewrite date in this form 2015-12-21.")
                return

        if len(ft) < 1:  # Default value, 在最后一行结束（Python下标从0开始，故最后一行index比excel小1）
            enddate = maxdate
        else:
            try:
                etmp = datetime.datetime.strptime(ft, "%Y-%m-%d")
                if etmp >= maxdate:
                    enddate = maxdate
                else:
                    enddate = etmp
            except:
                print("Wrong end date format, please rewrite date in this form 2015-12-21.")
                return

        #  更新参数
        self.startdate = startdate
        self.enddate = enddate
        return [startdate, enddate]


    def GetDep( self, Dep ):
        """
        function: 根据用户输入，确定所选部门。
        :param Dep: 用户输入的部门名称, Default value为空，即班表中的所有部门名称。
        :return: Dep，用户所选部门；0表示无法识别，Please re-enter department name.
        """

        #  当Dep合法或为Default value时。
        if Dep in self.DEPT_TYPE or len(Dep) < 1:
            return Dep
        #  否则返回0，表示无法识别。
        else:
            return 0


    def GetClassName( self, cyconsts ):
        """
        function: 根据用户输入，确定用户所选课程类型。
        :param cyconsts: 用户输入的课程类型，Default value为空，即班表中的所有课程类型。
        :return: cyconsts， 用户所选课程类型；0表示无法识别，Please re-enter class type name.
        """

        #  当Dep合法或为Default value时。
        if cyconsts in self.CLASS_TYPE.keys() or len(cyconsts) < 1:
            return cyconsts
        #  否则返回0，表示无法识别。
        else:
            return 0

    def isinClassType(self, CLASS_TYPE, Classprogram, CurrentClass, StudentNum):
        """
        function: 检查班表中的当前课程是否满足用户所选的约束条件。
        :param
        CLASS_TYPE: 包含所有的课程类型，dict{课程项目名(str):[课程包含关键字](list)}；
        Classprogram: 课程类型名(str)，一般表示用户所选的课程类型；
        CurrentClass: 当前课程名(str)；
        StudentNum: 学生人数(str)；
        :return: True if当前课程CurrentClass，属于（用户选择的）课程类型Classprogram; False otherwise。
        """

        flag = False
        if "VIP" in Classprogram:  # VIP课一定是1对1 或 6人
            if StudentNum != "1" and StudentNum != "6":
                return False
        if "班级" in Classprogram:  # 班级项目一定不是 1对1 或 6人的
            if StudentNum == "1" or StudentNum == "6":
                return False
        for i in range(len(CLASS_TYPE[Classprogram])):
            if CLASS_TYPE[Classprogram][i] in CurrentClass:
                flag = True
        return flag


    def GetStuNum(self, current_num, class_name):
        """
        function: 根据班表中的2个参数，计算该班级的学生数。
        :param current_num:班表中的当前人数，即导出班表的时刻该班级中的学生数。
                             如果此时该班已结课或退课，则为0，否则则显示真实学生数。
        :param class_name:班表中的班级名称，当无法获得真实学生数时，可通过班级名称中的关键字估算学生数。
        :return: 该班级中的学生数量。
        """

        #  如果当前人数不为零，则为真实学生数，直接返回该值。
        if current_num != 0:
            return current_num
        #  当前人数为零时，则需要从班级名称中寻找关键字。
        if "二" in class_name:
            return 2
        elif "三" in class_name:
            return 3
        elif "1" in class_name:
            return 1
        elif "8" in class_name:
            return 8
        elif "15" in class_name:
            return 15
        elif ("基础" or "强化" or "全程") in class_name:
            return 8
        else:  # "定制" in class_name
            return 15


    def FilterClsTable( self, st, ft, Dep, cyconsts, teachers, output_mode ):
        """
        function: 根据参数，计算一段时间内，某部门（或全部），某课程（或全部）的产能，与各教师在其中的产能列表。
        :param st & ft: 起始 & 终止时间范围；Dep: 项目部门； cyconsts: 课程类型； teachers: 教师姓名。
               output_mode: 是否输出txt和excel文件。
        :return: summary_text：梳理统计后的综述，teacherdict（教师产能列表）：[姓名，总产能，总课时，产能/课时，产能/月，工作月份]。
                 输出2个txt文档：Summary.txt 和 Classes_table.txt，即resclstab的内容；
                 输出1个excel文档：Teacher_List.xlsx，即teacherdict的内容。
        """

        #  载入类属性值。
        table = self.table
        rows = self.rows
        cols = self.cols

        #  STEP1. 根据用户输入的时间范围，找出table中的index。
        [startdate, enddate] = self.GetDateRange( st, ft )

        #  STEP2. 用户输入要查询的项目部门。
        Dep = self.GetDep( Dep )

        #  STEP3. 用户输入要查询的课程类型。
        cyconsts = self.GetClassName( cyconsts )

        #  STEP4. 根据前几步的约束条件，重新整理 & 统计班表。
        #  (1).以开班编号为primary key来扫描并统计班表。
        #  用resclstab和classes存放计算结果。
        #  赋初值：
        resclstab = []  # Result Class Table.
        #  用于计算的classes dictionary.
        #  班级编号为key，value为一个dict{该班级的所有老师名：其对应的上课时长}。
        classes = {}
        header = table.row_values(0)  # 第一行为表头。
        for i in range( 1, rows ):
            r = table.row_values(i)
            curdep = r[self.DEPT_POS]  # 当前部门。
            curcls = r[self.CLASS_POS]  # 当前课程种类。
            curdate = xlrd.xldate_as_datetime(r[self.CLASS_DATE_POS], 0)  # 当前日期。
            curteacher = r[self.TEACHERS_POS]  # 当前教师姓名。

            #  Sentinels
            #  Teacher name Constraint.
            #  例如teachers=“刘莹”，curteacher=“刘莹-雅思阅读”，则为合法数据。
            if teachers not in curteacher and len(teachers) >= 1:
                continue
            #  Date Constraint.
            if curdate < startdate or curdate > enddate:
                continue
            #  Department Constraint.
            if len(Dep) >= 1:  # 没有使用default的情况。
                if Dep not in curdep:
                    continue
            #  Class Type Constraint.
            if len(cyconsts) >= 1:  # 没有使用default的情况。
                flag = self.isinClassType(self.CLASS_TYPE, cyconsts, curcls, r[self.STU_NUM_POS])
                if not flag:
                    continue
            #  排除自习课、辅导课、模考等。
            if len(r[self.TEACHERS_POS]) == 0:
                continue
            #  排除没学费的课程。
            if r[self.FEE_POS] == "#N/A":
                continue

            curclstime = float(r[self.CLASS_TIME_POS])  # 当前班级所用学时（单位：分钟）。
            totalclstime = r[self.CLASS_TIMESUM_POS]  # 该班级的总学时（单位：分钟）。
            clsnum = r[self.CLASS_NUM]  # 当前课号。
            cur_stunum = self.GetStuNum(int(r[self.STU_NUM_POS]), r[self.CLASS_POS])

            #  当前课号第一次出现时：
            if clsnum not in classes.keys():
                t4tc = {curteacher: curclstime}  # teacher for this class.
                #  更新classes dictionary.
                classes[clsnum] = t4tc

                #  更新Result Class Table.
                #  将课程总时长（str）转换为float。
                try:
                    studyhours = float(totalclstime)
                    if studyhours == 0:  # 班表中总时长为0的情况。
                        studyhours = None
                except:
                    studyhours = None  # 班表中总时长为空的情况。

                tmp = [curdep, curcls, clsnum, float(r[self.FEE_POS]), studyhours, cur_stunum]
                resclstab.append(tmp)
            else:
                #  当前课号已存在于classes中的情况：
                if curteacher in classes[clsnum].keys():  # 当前教师已存在于该课号中。
                    classes[clsnum][curteacher] += curclstime
                else:  # 当前教师不存在于该课号中。
                    classes[clsnum][curteacher] = curclstime

        #  (2).以教师为primary key，整理&修改classes和resclstab。
        totalclsn = len(resclstab)  # Total classes number.
        totalfee = 0  # 总业绩初值。
        teacherdict = {}  # 参与教师dict初值。

        for i in range(totalclsn):
            clsnum = resclstab[i][2]
            #  计算该班的业绩。
            #  resclstab[i][3]表示该班级每位学生的学费。
            #  resclstab[i][4]表示该班级的总时长（单位：分钟）。
            #  resclstab[i][5]表示该班级的学生数量。

            fee = resclstab[i][3] * resclstab[i][5]
            #  该班级的业绩加入总业绩。
            totalfee += fee
            tct = resclstab[i][4]  # Total Class Time.

            #  不定时课程的情况，则将所有老师的课时叠加作为总课时。
            if tct is None:
                tct = 0
                for t in classes[clsnum].keys():
                    tct += classes[clsnum][t]

            #  更新该课程中每位老师的课数、课时贡献百分比 和 对应的完成业绩。
            for t in classes[clsnum].keys():
                #  首次统计到该老师时，将其姓名加入teacherdict。
                if t not in teacherdict.keys():
                    #  教师姓名t：总产能，总课时，产能/小时，产能/月，最早上课日期，最晚上课日期, month_rate。
                    teacherdict[t] = [0, 0, 0, 0, None, None, None]
                time = classes[clsnum][t]  # t老师在该班级上了time分钟的课。
                p = time / tct
                #  更新classes，存入t老师在该班级上了times课，占所有课次的比例，以及对应该班级学费的业绩。
                classes[clsnum][t] = [time, round(p, 5), round(p * fee, 2)]
            #  更新resclstab中的classes信息。
            resclstab[i].append(classes[clsnum])
        # 参与教师总数。
        totaltn = len(teacherdict.keys())

        #  STEP5. 计算教师产值列表。
        #  按每个班级中的每个教师的产能，叠加到teacherdict中该教师的总产能。
        for c in classes.keys():
            for t in classes[c].keys():
                individ_time = classes[c][t][0]  # 个人时长。
                individ_performance = classes[c][t][2]  # 个人业绩。
                teacherdict[t][0] += individ_performance  # 累积个人业绩。
                teacherdict[t][1] += round(individ_time/60, 2)  # 累积个人课时（单位：小时）。

        # 更新产能课时比。
        for t in teacherdict.keys():
            teacherdict[t][0] = round(teacherdict[t][0], 2)
            teacherdict[t][2] = round(teacherdict[t][0] / teacherdict[t][1], 2)  # 更新每学时产能。

            #  在班表中查找t老师上课的最早日期和最晚日期。
            #  对于最早和最晚日期赋初值。
            earliest = datetime.datetime.strptime("2200-01-01", "%Y-%m-%d")
            latest = datetime.datetime.strptime("1880-01-01", "%Y-%m-%d")
            for i in range( 1, rows ):
                r = table.row_values(i)
                curdate = xlrd.xldate_as_datetime(r[self.CLASS_DATE_POS], 0)  # 当前日期。
                curteacher = r[self.TEACHERS_POS]  # 当前教师姓名。
                curdep = r[self.DEPT_POS]  # 当前部门。
                curcls = r[self.CLASS_POS]  # 当前课程种类。
                #  Sentinels.
                #  不是t老师则跳过。
                if curteacher != t:
                    continue
                #  Date Constraint.
                if curdate < startdate or curdate > enddate:
                    continue
                #  Department Constraint.
                if len(Dep) >= 1:  # 没有使用default的情况。
                    if Dep not in curdep:
                        continue
                #  Class Type Constraint.
                if len(cyconsts) >= 1:  # 没有使用default的情况。
                    flag = self.isinClassType(self.CLASS_TYPE, cyconsts, curcls, r[self.STU_NUM_POS])
                    if not flag:
                        continue
                #  排除自习课、辅导课、模考等。
                if len(r[self.TEACHERS_POS]) == 0:
                    continue
                #  排除没学费的课程。
                if r[self.FEE_POS] == "#N/A":
                    continue
                if curdate < earliest:
                    earliest = curdate
                if curdate > latest:
                    latest = curdate
            # month_rate为[startdate, enddate]期间带课的月份。
            month_rate = round((latest - earliest).days / 30, 2)
            teacherdict[t][3] = round(teacherdict[t][0] / month_rate, 2)  #  更新每月产能。
            teacherdict[t][4] = earliest
            teacherdict[t][5] = latest
            teacherdict[t][6] = month_rate

        #  排序教师产值.
        #  对teacherdict（dict）的value排序。
        teacherdict = sorted(teacherdict.items(), key=lambda item: item[1], reverse=True)

        #  STEP6.创建返回值output_text和teacherlist。
        if len(Dep) < 1:
            Dep = "全部项目部"
        if len(cyconsts) < 1:
            cyconsts = "全部"

        #  datetime转字符串。
        start = startdate.strftime("%Y-%m-%d %H:%M:%S")
        end = enddate.strftime("%Y-%m-%d %H:%M:%S")

        output_text = ("您选择了 " + start + " 到 " + end + "\n"
                       + Dep + "\n"
                       + cyconsts + "课程项目\n"
                       + "共开班：" + str(totalclsn) + "节\n"
                       + "共" + str(totaltn) + "位教师参与教学\n"
                       + "完成业绩：" + str(totalfee) + "元\n"
                       + "********************************************"
                       + "\n"
                       + "\n")
        #  赋初值。
        teacherlist = []
        perform_mean = 0
        hours_mean = 0
        pperh_mean = 0
        pperm_mean = 0
        month_mean = 0
        for i in range(totaltn):
            #  复杂的dict无法直接输出到excel，先转换成list。
            teacher_name = teacherdict[i][0]  # 教师姓名。
            perform_sum = teacherdict[i][1][0]  # 总产能。
            hours_sum = teacherdict[i][1][1]  # 总课时。
            pperh = teacherdict[i][1][2]  # 产能/课时。
            pperm = teacherdict[i][1][3]  # 产能/月。
            earliest_str = teacherdict[i][1][4].strftime("%Y-%m-%d")  # 带课的最早日期。
            latest_str = teacherdict[i][1][5].strftime("%Y-%m-%d")  # 带课的最晚日期。
            month_rate = teacherdict[i][1][6]

            perform_mean += perform_sum
            hours_mean += hours_sum
            pperh_mean += pperh
            pperm_mean += pperm
            month_mean += month_rate
            teacherlist.append([teacher_name,perform_sum,hours_sum,pperh,pperm,earliest_str,latest_str, month_rate])
        #  teachlist不为空才进行总结。
        if totaltn > 0:
            teacherlist.append([" ", " ", " ", " ", "", " ", " ", " "])
            teacherlist.append([" ", "平均产能", "平均课时", "平均产能/课时", "平均产能/月", " ", " ", "平均授课时间（月）"])
            teacherlist.append([" ", round(perform_mean/totaltn), round(hours_mean/totaltn), round(pperh_mean/totaltn), round(pperm_mean/totaltn), " ", " ", round(month_mean/totaltn)])

        #  STEP7. 创建3个输出文件。
        #  输出结果文件1：“Summary.txt”.
        if output_mode:
            f = open("Summary.txt", "a+")  # 如果存在该文件则在末尾添加，否则就创建该文件。
            f.write(output_text)
            f.close()

            #  输出结果文件2：“Classes_Table.txt”.
            f = open("Classes_Table.txt", "a+")  # 如果存在该文件则在末尾添加，否则就创建该文件。
            for i in range(len(resclstab)):
                f.writelines(str(resclstab[i]) + "\n")
            f.write(("********************************************"
                     + "\n"
                     + "\n"))
            f.close()

            #  输出结果文件3：“Teacher_List.xlsx”.
            #  获取当前路径，并查看Teacher_List.xlsx是否存在。
            cwd = os.getcwd()
            #  文件存在，则新建sheet。
            if os.path.exists(cwd + "\Teacher_List.xlsx"):
                wb = load_workbook("Teacher_List.xlsx")
                ws = wb.create_sheet()  # 默认插在末尾。
            #  文件不存在则新建。
            else:
                #  在内存中创建一个workbook对象，而自动创建一个worksheet。
                wb = Workbook()
                #  获取当前活跃的worksheet，默认就是第一个worksheet。
                ws = wb.active
            #  写入excel。
            ws.append(["教师姓名","总产能","总课时","产能/课时","产能/月","首节课日期","末节课日期","带课时间（月）"])
            for i in range(len(teacherlist)):
                ws.append(teacherlist[i])
            #  保存excel。
            wb.save("Teacher_List.xlsx")
        return teacherlist, output_text


def CalcuAll( CLASS_SCHEDULE_FILE, output_mode ):
    """
    function: 试算2018.06.01-2019.05-31，所有部门，所有老师的产能。
    :param CLASS_SCHEDULE_FILE: 班表名称。
    :return: 输出2个txt, 1个excel文档。
    """

    #  定义类函数所需的parameters。
    st = str()
    ft = str()
    Dep = str()
    cyconsts = str()
    teachers = str()

    #  创建Teacher's performance Trackers实例。
    fy19 = TpTrackers( CLASS_SCHEDULE_FILE )
    #  计算Teacher's performance。
    return fy19.FilterClsTable(st, ft, Dep, cyconsts, teachers, output_mode)


def PerfomafterEntry( CLASS_SCHEDULE_FILE, ENTRYTIME_FILE, timewindow, timelimit ):
    """
    function: 试算每位老师入职后timewindow时间内的产能。
    :param CLASS_SCHEDULE_FILE: 班表名称。
    :param ENTRYTIME_FILE: 老师入职时间表。
    :param timewindow: 要计算的入职后的时间范围（单位：月）。
    :param timelimit: 要计算的开始带课后的时间范围（单位：月）。
    :return entryperfom: 每位老师入职后timewindow时间内的产能列表。
            [教师姓名，总产能，总课时，产能/课时，产能/月，首节课日期, 末节课日期, 带课时间（月），入职时间]
    """

    #  打开入职时间表。
    data1 = xlrd.open_workbook( ENTRYTIME_FILE )
    nametable = data1.sheets()[0]
    m = nametable.nrows

    #  定义常量。
    TEACHER_POS = 2  #  教师姓名在excel中的第2列（从第0列开始）。
    ENTRY_DATE_POS = 3  #  入职日期在excel中的第2列（从第0列开始）。
    NOW_DATE = datetime.datetime.strptime("2019-05-21", "%Y-%m-%d")  #  入职至今，即“今”表示的日期。

    #  创建Teacher's performance Trackers实例。
    fy19 = TpTrackers( CLASS_SCHEDULE_FILE )

    #  定义function所需的其他参数。
    Dep = str()
    cyconsts = str()

    #  对结果赋初值。
    entryperform =[]
    #  对均值赋初值。
    perform_mean = 0
    hours_mean = 0
    pperh_mean = 0
    pperm_mean = 0
    count = 0  # 新教师人数。
    header = nametable.row_values(0)  # 第一行为表头。
    for i in range(1, m):
        r = nametable.row_values(i)
        teachers = r[TEACHER_POS]    # 读取教师姓名。

        #  读取入职时间。
        entry_date = xlrd.xldate_as_datetime(r[ENTRY_DATE_POS], 0)
        end_date = entry_date + datetime.timedelta(days=(int(timewindow) * 30))

        #  datetime转str。
        entry_date_str = entry_date.strftime("%Y-%m-%d")
        end_date_str = end_date.strftime("%Y-%m-%d")

        #  计算Teacher's performance。
        [tmp, detail] = fy19.FilterClsTable(entry_date_str, end_date_str, Dep, cyconsts, teachers, False)

        #  跳过空值的计算结果。
        if len(tmp) < 1:
            continue

        delta = (NOW_DATE - entry_date).days
        #  筛选入职不超过15个月，带课不超过12个月的新老师。
        if delta < ((timewindow - 12) * 31 + 365) and tmp[0][7] <= 12:
            tmp = list(tmp[0])  # 从[[]]修改为[]
            tmp.append(entry_date.strftime("%Y-%m-%d"))
            entryperform.append(tmp)
            count += 1
            perform_mean += tmp[1]
            hours_mean += tmp[2]
            pperh_mean += tmp[3]
            pperm_mean += tmp[4]

    entryperform.append([" ", " ", " ", " ", " ", " ", " ", " ", " "])
    entryperform.append([" ", "平均产能", "平均课时", "平均产能/课时", "平均产能/月", " ", " ", " "])
    entryperform.append([" ", round(perform_mean / count), round(hours_mean / count), round(pperh_mean / count),
               round(pperm_mean / count), " ", " ", " "])

    return entryperform


#------------------------------------------------------------------
#  Script:
#  定义function的parameter。
CLASS_SCHEDULE_FILE = "FY19国外配课表.xlsx"
ENTRYTIME_FILE = "【国外部】教师名单-3.7.xlsx"
timewindow = 15
timelimit = 12
CASE = 1

if CASE == 1:
    #  test1. 试算2018.06.01-2019.05-31，所有部门，所有老师的产能。
    CalcuAll( CLASS_SCHEDULE_FILE, True )
elif CASE == 2:
    # test3. 试算每位老师入职不超过timewindow个月，带课不超过timelimit个月的产能表现。
    entryperform = PerfomafterEntry(CLASS_SCHEDULE_FILE, ENTRYTIME_FILE, timewindow, timelimit)
    #  输出到excel。
    #  在内存中创建一个workbook对象，而自动创建一个worksheet。
    wb = Workbook()
    #  获取当前活跃的worksheet，默认就是第一个worksheet。
    ws = wb.active
    ws.append(["教师姓名","总产能","总课时","产能/课时","产能/月","首节课日期","末节课日期","带课时间（月）","入职日期"])
    for i in range(len(entryperform)):
        ws.append(entryperform[i])
    #  保存excel。
    wb.save("Performance_after_entry.xlsx")


